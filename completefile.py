from PIL import Image
from skimage.color import rgb2lab
from scipy.spatial import KDTree
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from colors import origamiColors, vliegerColor, sixInchTuttleColors
import tkinter as tk
from tkinter import filedialog
import os


def hex_to_rgba(hex_string):
    hex_string = hex_string.lstrip('#')
    return tuple(int(hex_string[i:i+2], 16) for i in (0, 2, 4, 6))


def hex_to_rgb_arr(hex_str):
    hex_str = hex_str.lstrip('#')[:6]
    return [int(hex_str[i:i+2], 16) / 255.0 for i in range(0, 6, 2)]


def pixelate_to_size(image_path, scale_factor):
    image = Image.open(image_path)
    width, height = image.size
    pixelated_image = image.resize(
        (width // scale_factor, height // scale_factor), 
        resample=Image.NEAREST
    )
    return pixelated_image


def match_lab(input_hex, palette_hex):
    input_rgb = np.array([hex_to_rgb_arr(h) for h in input_hex])
    palette_rgb = np.array([hex_to_rgb_arr(h) for h in palette_hex])
    input_lab = rgb2lab(input_rgb.reshape(-1, 1, 3)).reshape(-1, 3)
    palette_lab = rgb2lab(palette_rgb.reshape(-1, 1, 3)).reshape(-1, 3)
    tree = KDTree(palette_lab)
    _, indices = tree.query(input_lab)
    return [h[:7] for h in (palette_hex[i] for i in indices)]


def map_image_to_palette_indices(image_path, kdtree):
    image = Image.open(image_path).convert("RGBA")
    img_array = np.array(image)
    h, w = img_array.shape[:2]
    rgb_pixels = img_array[:, :, :3].reshape(-1, 3)
    _, indices = kdtree.query(rgb_pixels)
    indices_2d = indices.reshape(h, w)
    return indices_2d


def save_mapped_image(indices_2d, palette_rgba, output_path):
    h, w = indices_2d.shape
    img_data = np.zeros((h, w, 4), dtype=np.uint8)
    for idx, color in enumerate(palette_rgba):
        img_data[indices_2d == idx] = color
    img = Image.fromarray(img_data, mode="RGBA")
    img.save(output_path)
    print(f"Saved mapped image to: {output_path}")


def save_indices_to_excel(indices_2d, palette_rgba, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Color Map"
    for i, color in enumerate(palette_rgba):
        hex_color = f'{color[0]:02X}{color[1]:02X}{color[2]:02X}'
        cell = ws.cell(row=1, column=i+1, value=i)
        cell.fill = PatternFill(
            start_color=hex_color, 
            end_color=hex_color, 
            fill_type="solid"
        )
    counts = np.bincount(indices_2d.flatten(), minlength=len(palette_rgba))
    for i, count in enumerate(counts):
        ws.cell(row=2, column=i+1, value=int(count))
    for i in range(indices_2d.shape[0]):
        for j in range(indices_2d.shape[1]):
            ws.cell(row=i+4, column=j+1, value=int(indices_2d[i, j]))
    wb.save(output_path)
    print(f"Saved Excel sheet to: {output_path}")


def select_image_file():
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(
        title="Select an image file",
        filetypes=[
            ("Image files", "*.png *.jpg *.jpeg *.gif *.bmp *.tiff"),
            ("All files", "*.*")
        ]
    )
    return file_path


def select_save_location(default_name):
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.asksaveasfilename(
        title="Save file as",
        defaultextension=os.path.splitext(default_name)[1],
        filetypes=[
            ("PNG files", "*.png"),
            ("Excel files", "*.xlsx"),
            ("All files", "*.*")
        ],
        initialfile=default_name
    )
    return file_path


def process_image(input_image_path, output_image_path, output_excel_path, 
                  hex_palette, scale_factor):
    print("Step 1: Pixelating image...")
    pixelated = pixelate_to_size(input_image_path, scale_factor)
    temp_path = "temp_pixelated.png"
    pixelated.save(temp_path)
    
    print("Step 2: Preparing color palette...")
    rgba_palette = np.array([hex_to_rgba(h) for h in hex_palette], dtype=np.uint8)
    rgb_palette = rgba_palette[:, :3]
    kdtree = KDTree(rgb_palette)
    
    print("Step 3: Mapping colors to palette...")
    indices = map_image_to_palette_indices(temp_path, kdtree)
    
    print("Step 4: Exporting results...")
    save_mapped_image(indices, rgba_palette, output_image_path)
    save_indices_to_excel(indices, rgba_palette, output_excel_path)
    
    print("\nProcessing complete!")


if __name__ == "__main__":
    hex_palette = vliegerColor()
    name_palette = "vliegerColor"
    print("Please select an image file...")
    
    input_image_path = select_image_file()
    
    if not input_image_path:
        print("No file selected. Exiting.")
    else:
        print(f"Selected: {input_image_path}")
        
        print("\nChoose save location for mapped image...")
        output_image_path = select_save_location(f"{input_image_path}_{name_palette}.png")
        
        if not output_image_path:
            print("No save location selected. Exiting.")
        else:
            print(f"Will save image to: {output_image_path}")
            
            print("\nChoose save location for Excel file...")
            output_excel_path = select_save_location(f"{input_image_path}_{name_palette}.xlsx")
            
            if not output_excel_path:
                print("No save location selected. Exiting.")
            else:
                print(f"Will save Excel to: {output_excel_path}")
                
                process_image(
                    input_image_path=input_image_path,
                    output_image_path=output_image_path,
                    output_excel_path=output_excel_path,
                    hex_palette=hex_palette,
                    scale_factor=30
                )